import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re


class ExcelObject:
    def __init__(self, excel_path=None, company_name=None, sheet_name=None, start_row=None, end_row=None, num=None):
        self.excel_path = excel_path
        self.company_name = company_name
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.end_row = end_row
        self.num = num


def find_row_range(excel_path=None, sheet_name=None) -> [int, int, int, int]:
    wb = load_workbook(excel_path)
    sheet_name = sheet_name
    sheet = wb[sheet_name]
    # print("B1单元格的值为：\n", sheet['B1'].value)
    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value == '1':
            start_row = cell.row
            # print("起始数据的行索引为：\n", start_row)
            break
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=False):
        cell = row[0]
        if cell.value is None:
            if cell.row == start_row:
                end_row = start_row
                num = 0
            else:
                end_row = cell.row - 1
                num = end_row - start_row + 1
            # print("末尾数据的行索引为：\n", end_row)
            break
    return [start_row, end_row, num]


def remove_parentheses(text):
    pattern = r'[\(\（][^()（）]*[\)\）]'
    result = re.sub(pattern, '', text)
    return result


work_path = r"C:\Users\ALGU1VY\Desktop\VASS_File_temp"
tar_excel_path = r"./Data/Fee in total_2024Q3.xlsx"

print("Please close the corresponding file that needs to be operated when executing this program!!!")
while True:
    season_name = input("Please enter the exam quarter for the desired operation (e.g. 202407):\n")
    season_path = os.path.join(work_path, season_name)
    try:
        if os.path.exists(season_path) and os.path.isdir(season_path):
            break
        else:
            raise FileNotFoundError
    except FileNotFoundError:
        print("The exam quarter does not exist, please re-enter!")

excel_extensions = ['.xls', '.xlsx']
excel_objects = {}
sheet_name_1 = 'Allg.'
sheet_name_2 = 'Company Invoice Infor'
i = 0
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

tar_excel_object = ExcelObject(excel_path=tar_excel_path, sheet_name='工作表1', start_row=3)

for item in os.listdir(season_path):
    item_path = os.path.join(season_path, item)
    if os.path.isdir(item_path):
        print(item)
        # print(company_path_list)
        for filename in os.listdir(item_path):
            if any(filename.endswith(ext) for ext in excel_extensions):
                # print(filename)
                excel_path = os.path.join(item_path, filename)
                print(excel_path)
                excel_objects[f'excel_objects_{i}'] = ExcelObject(excel_path=excel_path, company_name=item)
                i += 1

company_num = len(excel_objects)
n = 0
# print(f'共有 {len(excel_objects)} 个公司。')
for _, excel_object in excel_objects.items():
    excel_object.sheet_name = sheet_name_1
    excel_object.start_row, excel_object.end_row, excel_object.num = (
        find_row_range(excel_path=excel_object.excel_path, sheet_name=excel_object.sheet_name))

    # print(vars(excel_object))
    wb_tar = load_workbook(tar_excel_object.excel_path)
    sheet_tar = wb_tar[tar_excel_object.sheet_name]

    if "考生公司与报名公司不一致" in excel_object.company_name:
        sheet_tar.cell(row=tar_excel_object.start_row + n, column=2).fill = yellow_fill
    elif "保留名额" in excel_object.company_name:
        sheet_tar.cell(row=tar_excel_object.start_row + n, column=2).fill = blue_fill
    elif "未付款" in excel_object.company_name:
        sheet_tar.cell(row=tar_excel_object.start_row + n, column=2).fill = red_fill
    elif "不安排考试" in excel_object.company_name:
        continue

    company_rename = remove_parentheses(excel_object.company_name)
    sheet_tar.cell(row=tar_excel_object.start_row+n, column=2, value=company_rename)
    sheet_tar.cell(row=tar_excel_object.start_row+n, column=4, value=excel_object.num)
    n += 1

    wb_tar.save(tar_excel_path)

print("Processing completed!")
