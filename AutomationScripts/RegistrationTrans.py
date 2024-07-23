from openpyxl import load_workbook
import os
from openpyxl.styles import PatternFill, Border, Side


class SheetObject:
    def __init__(self, excel_path=None, company_name=None, sheet_name=None, start_row=None, end_row=None, num=None):
        self.excel_path = excel_path
        self.company_name = company_name
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.end_row = end_row
        self.num = num


def ori_row_range(excel_path=None, sheet_name=None) -> [int, int, int]:
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    # print("B1单元格的值为：\n", sheet['B1'].value)
    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value == '1':
            start_row = cell.row
            # print("起始数据的行索引为：\n", start_row)
            break
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=False):
        cell = row[0]
        if cell.value is None:
            if cell.row == start_row:
                end_row = start_row
                num = 0
            else:
                end_row = cell.row - 1
                num = end_row - start_row + 1
            # print("末尾数据的行索引为：\n", end_row)
            break
    return [start_row, end_row, num]


def tar_row_range(excel_path=None, sheet_name=None) -> int:
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    count = 0
    for row_cell in sheet.iter_rows(min_row=134, min_col=3, max_col=3, values_only=False):
        # print(row_cell[0])
        if row_cell[0].value is not None and count == 2:
            count = 0
        elif row_cell[0].value is None:
            count += 1
        if count == 3:
            start_row = row_cell[0].row
            print(start_row, type(tar_sheet_object.start_row))
            return start_row


ori_excel_path = r".\Data\Registration form_all_2024Q3_test.xlsx"
tar_excel_path = r".\Data\Registration form_all_V03.xlsx"
sheet_name = "Allg."
light_yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ori_sheet_object = SheetObject(excel_path=ori_excel_path, sheet_name='Allg.')
tar_sheet_object = SheetObject(excel_path=tar_excel_path, sheet_name='Allg.')

ori_sheet_object.start_row, ori_sheet_object.end_row, ori_sheet_object.num = (
        ori_row_range(excel_path=ori_sheet_object.excel_path, sheet_name=ori_sheet_object.sheet_name))
tar_sheet_object.start_row = tar_row_range(excel_path=tar_excel_path, sheet_name=tar_sheet_object.sheet_name)
# print(ori_sheet_object.start_row)
print(tar_sheet_object.start_row)

wb_ori = load_workbook(ori_excel_path)
wb_tar = load_workbook(tar_excel_path)

sheet_ori = wb_ori[sheet_name]
sheet_tar = wb_tar[sheet_name]
order_start_num = sheet_tar.cell(row=int(tar_sheet_object.start_row)-3, column=1).value
print(order_start_num)

for i in range(ori_sheet_object.num):
    sheet_tar.cell(row=tar_sheet_object.start_row+i, column=1, value=order_start_num+1+i)
    sheet_tar.cell(row=tar_sheet_object.start_row+i, column=1).fill = light_yellow_fill
    for j in range(7):
        ori_value = sheet_ori.cell(row=ori_sheet_object.start_row+i, column=j+2).value
        sheet_tar.cell(row=tar_sheet_object.start_row+i, column=j+2, value=ori_value)
        sheet_tar.cell(row=tar_sheet_object.start_row+i, column=j+2).border = thin_border

wb_tar.save(tar_excel_path)
print("Processing completed!")
