from openpyxl import load_workbook, workbook
import os


class SheetObject:
    def __init__(self, excel_path=None, company_name=None, sheet_name=None, start_row=None, end_row=None, num=None):
        self.excel_path = excel_path
        self.company_name = company_name
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.end_row = end_row
        self.num = num


def find_row_range(excel_path=None, sheet_name=None) -> [int, int, int, int]:
    wb = load_workbook(excel_path)
    sheet_name = sheet_name
    sheet = wb[sheet_name]
    # print("B1单元格的值为：\n", sheet['B1'].value)
    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value == '1':
            start_row = cell.row
            # print("起始数据的行索引为：\n", start_row)
            break
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=False):
        cell = row[0]
        if cell.value is None:
            if cell.row == start_row:
                end_row = start_row
                num = 0
            else:
                end_row = cell.row - 1
                num = end_row - start_row + 1
            # print("末尾数据的行索引为：\n", end_row)
            break
    return [start_row, end_row, num]


ptc_path = r"C:\Users\ALGU1VY\Desktop\VASS_File_temp"
print("Please close the corresponding file that needs to be operated when executing this program!!!")
test_season = input("Please enter the exam quarter for the desired operation (e.g. 202407):\n")
season_path = os.path.join(ptc_path, test_season)
target_sheet_path = r"./Data/Registration form_all_2024Q3_test.xlsx"
assert os.path.exists(season_path), f"File:'{season_path}' doesn't exist, please input it again!"
assert os.path.exists(target_sheet_path), f"File:'{target_sheet_path}' doesn't exist, please input it again!"

excel_extensions = ['.xls', '.xlsx']
company_path_list = []
company_excel_path = []
sheet_objects_allg = {}
sheet_objects_invoice = {}
sheet_name_1 = 'Allg.'
sheet_name_2 = 'Company Invoice Infor'
company_num = 0
# print(season_path)

for item in os.listdir(season_path):
    item_path = os.path.join(season_path, item)
    if os.path.isdir(item_path):
        print(item)  # 公司名
        company_path_list.append(item_path)
# print(company_path_list)
        for filename in os.listdir(item_path):
            if any(filename.endswith(ext) for ext in excel_extensions):
                # print(filename)
                excel_path = os.path.join(item_path, filename)
                company_excel_path.append(excel_path)
                sheet_objects_allg[f'sheet_object_{company_num}'] = SheetObject(excel_path=excel_path, company_name=item)
                sheet_objects_invoice[f'sheet_object_{company_num}'] = SheetObject(excel_path=excel_path, company_name=item)
                company_num += 1
excel_num = len(company_excel_path)

# print(excel_num)
allg_sheet_object = SheetObject(excel_path=target_sheet_path, sheet_name='Allg.')
invoice_sheet_object = SheetObject(excel_path=target_sheet_path, sheet_name='Company Invoice Infor')
# print(sheet_objects_allg)

# print(company_excel_path)
allg_sheet_object.start_row, allg_sheet_object.end_row, allg_sheet_object.num = (
        find_row_range(excel_path=allg_sheet_object.excel_path, sheet_name=allg_sheet_object.sheet_name))
invoice_sheet_object.start_row, invoice_sheet_object.end_row, invoice_sheet_object.num = (
        find_row_range(excel_path=invoice_sheet_object.excel_path, sheet_name=invoice_sheet_object.sheet_name))
# print(vars(allg_sheet_object))
# print(vars(invoice_sheet_object))
for _, sheet_object in sheet_objects_allg.items():
    # print(sheet_object.excel_path)
    if "不安排考试" in sheet_object.company_name:
        continue
    sheet_object.sheet_name = sheet_name_1
    sheet_object.start_row, sheet_object.end_row, sheet_object.num = (
        find_row_range(excel_path=sheet_object.excel_path, sheet_name=sheet_object.sheet_name))

    wb_ori = load_workbook(sheet_object.excel_path)
    wb_tar = load_workbook(target_sheet_path)

    sheet_ori_allg = wb_ori['Allg.']
    sheet_tar_allg = wb_tar['Allg.']

    # print(vars(sheet_object))
    # input("pause!")
    for i in range(sheet_object.num):
        for j in range(7):
            ori_value_allg = sheet_ori_allg.cell(row=sheet_object.start_row+i, column=2+j).value
            sheet_tar_allg.cell(row=allg_sheet_object.start_row+i, column=2+j, value=ori_value_allg)

    allg_sheet_object.start_row += sheet_object.num
    wb_tar.save(target_sheet_path)

for _, sheet_object in sheet_objects_invoice.items():
    # print(sheet_object.excel_path)
    if "不安排考试" in sheet_object.company_name:
        continue
    sheet_object.sheet_name = sheet_name_2
    sheet_object.start_row, sheet_object.end_row, sheet_object.num = (
        find_row_range(excel_path=sheet_object.excel_path, sheet_name=sheet_object.sheet_name))

    wb_ori = load_workbook(sheet_object.excel_path)
    wb_tar = load_workbook(target_sheet_path)

    sheet_ori_invoice = wb_ori['Company Invoice Infor']
    sheet_tar_invoice = wb_tar['Company Invoice Infor']

    # print(vars(sheet_object))
    # input("pause!")
    for i in range(sheet_object.num):
        for j in range(10):
            ori_value_invoice = sheet_ori_invoice.cell(row=sheet_object.start_row+i, column=2+j).value
            sheet_tar_invoice.cell(row=invoice_sheet_object.start_row+i, column=2+j, value=ori_value_invoice)

    invoice_sheet_object.start_row += sheet_object.num
    wb_tar.save(target_sheet_path)

print("Processing completed!")